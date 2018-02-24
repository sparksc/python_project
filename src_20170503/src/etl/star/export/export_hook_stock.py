# -*- coding:utf-8 -*- 
from nose.tools import eq_, raises, assert_true
from sqlalchemy.orm import aliased
import unittest
import time,os,random,sys
import DB2
from decimal import Decimal

import etl.base.util as util

import xlrd,xlwt

from openpyxl import Workbook,load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font,Color
import datetime

def run_export():
    data=query_detail()
    hook=query_hook()
    idx={}
    sheetlist={}
    booklist={}
    rowlist={}
    for i in range(0,len(data)):
        x=data[i]
        if x[0]+x[6] not in rowlist:
            [manage_code,htype]=hook.get(x[0]+x[6],['无','无'])
            taddr=util.get_cust_address(x[0])
            tlist=list(x)+list(taddr)+[manage_code,htype]
            rowlist[x[0]+x[6]]=tlist
        else:
            rowlist[x[0]+x[6]]=tlist
            print x[0]+x[6]
    print len(rowlist)        
    listlen={}
    for ii in rowlist:
        tlist=rowlist[ii]    
        if len(tlist) not in listlen:
            listlen[len(tlist)]=1
        else:    
            listlen[len(tlist)]=listlen[len(tlist)]+1
        sno=tlist[6]
        if sno not in booklist:
            booklist[sno]=xlwt.Workbook() 
            sheetlist[sno]=booklist[sno].add_sheet('sheet1')
            idx[sno]=1
        #tlist.pop(17)
        #tlist.pop(17)
        #tlist.pop(15)
        tlist.append(tlist[6])
        temp=tlist[15]
        tlist[15]=tlist[16]
        tlist[16]=temp
        tlist[17]=tlist[6]
        temp=tlist[17]
        tlist[17]=tlist[16]
        tlist[16]=temp
        #temp=tlist[15]
        #tlist[15]=tlist[16]
        #tlist[16]=temp
        #temp=tlist[14]
        #tlist[14]=tlist[15]
        #tlist[15]=temp
        for j in range(0,len(tlist)):
            sheetlist[sno].row(idx[sno]).write(j,str(tlist[j]).decode('utf-8'))
        idx[sno]=idx[sno]+1
    print listlen    
    for name in booklist:                
        sheetlist[name].row(0).write(0,u'客户内码')
        sheetlist[name].row(0).write(1,u'客户号')
        sheetlist[name].row(0).write(2,u'客户姓名')
        sheetlist[name].row(0).write(3,u'联系方式')
        sheetlist[name].row(0).write(4,u'业务类型')
        sheetlist[name].row(0).write(5,u'开始日期')
        sheetlist[name].row(0).write(6,u'开卡网点')
        sheetlist[name].row(0).write(7,u'卡号')
        sheetlist[name].row(0).write(8,u'帐号')
        sheetlist[name].row(0).write(9,u'客户姓名')
        sheetlist[name].row(0).write(10,u'证件类型')
        sheetlist[name].row(0).write(11,u'证件号码')
        sheetlist[name].row(0).write(12,u'开始日期')
        sheetlist[name].row(0).write(13,u'核心地址')
        sheetlist[name].row(0).write(14,u'信贷地址')
        sheetlist[name].row(0).write(15,u'认定方式')
        sheetlist[name].row(0).write(16,u'归属机构')
        sheetlist[name].row(0).write(17,u'归属柜员号')
        if not os.path.exists('tmp/'+name[:-1]+'0/第三方存管/'):
            os.makedirs('tmp/'+name[:-1]+'0/第三方存管/')
        booklist[name].save('tmp/'+name[:-1]+'0/第三方存管/'+name+'第三方存管.xls')
    #sumbook = xlwt.Workbook()    
    #sumsheet = sumbook.add_sheet('sheet1')   
    #idx=1
    #for sno in sumlist:
    #    tt=sno.split('_')
    #    tt.append("%.2f"%(sumlist[sno][0]/10000.0))
    #    tt.append("%.2f"%(sumlist[sno][1]/10000.0))
    #    tt.append("%.2f"%(sumlist[sno][2]/10000.0))
    #    print tt
    #    for i in range(0,len(tt)):
    #        sumsheet.row(idx).write(i,str(tt[i]))
    #    idx=idx+1    
    #sumbook.save("966120电子银行经理合计.xls")
def query_detail():
    try :
        db = util.DBConnect()
        q_sql="""
        SELECT a.CST_NO,a.CST_ID,cc.CST_NAME, MOBILE_NO, BUSI_TYPE, cc.OPEN_DATE,case a.OPEN_BRANCH_CODE when '966166' then '966163' else a.OPEN_BRANCH_CODE end, cc.CARD_NO, cc.ACCT_NO, ACCT_NAME, ID_TYPE, ID_NUMBER, THIRD_OPEN_DATE 
            FROM YDW.D_CUST_CONTRACT cc
            join D_ACCOUNT a on cc.ACCT_NO=a.ACCOUNT_NO and cc.ACCT_NO<>'无' --and a.OPEN_BRANCH_CODE in ('966120','966121','966122','966123','966125')
            where BUSI_TYPE='第三方存管' and THIRD_OPEN_DATE<=20160630
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchall()
        print len(row)
        return row
    finally :
        db.closeDB()

def query_hook():
    try :
        db = util.DBConnect()
        q_sql="""
        select distinct CUST_IN_NO,ORG_NO,MANAGER_NO from CUST_HOOK where TYP='存款' and HOOK_TYPE='管户'
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchone()
        rowdict={}
        while row:
            rowdict[row[0]+row[1]]=[row[2],'存款']
            row=db.cursor.fetchone()
        q_sql="""
        select distinct CUST_IN_NO,ORG_NO,MANAGER_NO from CUST_HOOK where TYP='贷款' and HOOK_TYPE='管户'
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchone()
        while row:
            rowdict[row[0]+row[1]]=[row[2],'贷款']
            row=db.cursor.fetchone()
        print len(rowdict)    
        return rowdict
    finally :
        db.closeDB()
if __name__=='__main__':
    arglen=len(sys.argv)
    print "begin:", str(datetime.datetime.now())
    run_export()
    print "end:", str(datetime.datetime.now())
    #if arglen==3:
    #    filename=sys.argv[1]
    #    hooktype=sys.argv[2]
    #    print filename,hooktype
    #    print "begin:", str(datetime.datetime.now())
    #    run_export(filename,hooktype)
    #    print "end:", str(datetime.datetime.now())
    #else:
    #    print "please input python export_hook.py [filename] [hooktype]"
