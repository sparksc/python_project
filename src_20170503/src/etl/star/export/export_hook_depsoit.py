# -*- coding:utf-8 -*- 
from nose.tools import eq_, raises, assert_true
from sqlalchemy.orm import aliased
import unittest
import time,os,random,sys
import DB2
from decimal import Decimal

import etl.base.util as util

import xlrd,xlwt

from openpyxl import Workbook,load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font,Color
import datetime

def run_export():
    dict2015=query_lrj()
    data=query_detail()
    idx={}
    sumlist={}
    sheetlist={}
    booklist={}
    for i in range(0,len(data)):
        temp=str(data[i][-1]).split(';')
        tempfc=str(data[i][-2])
        #print temp
        for x in temp:
            tt=[]
            tt.append(int(dict2015.get(data[i][0],0)))
            tt.append(int(data[i][12]))
            tt.append(tt[1]-tt[0])
            tt.append(tempfc)
            #print tt
            tlist=list(data[i][:-3])+tt+x.split('!')
            tlist[5]=int(tlist[5])/100.0
            sno=str(tlist[-3])+'_'+str(tlist[-6])
            if sno not in booklist:
                booklist[sno]=xlwt.Workbook() 
                sheetlist[sno]=booklist[sno].add_sheet('sheet1')
                sumlist[sno]=[]
                sumlist[sno].append(0)
                sumlist[sno].append(0)
                sumlist[sno].append(0)
                idx[sno]=1
            #else:    
            #if sno=="966115_9660518":print tlist
            for j in range(0,len(tlist)-1):
                #print str(data[i][j]).decode('utf-8')
                sheetlist[sno].row(idx[sno]).write(j,str(tlist[j+1]).decode('utf-8'))
                if int(tlist[-4])!=100:
                    print tlist
                if j==11:sumlist[sno][0]=sumlist[sno][0]+int(tlist[j+1])*int(tlist[-4])
                if j==12:sumlist[sno][1]=sumlist[sno][1]+int(tlist[j+1])*int(tlist[-4])
                if j==13:sumlist[sno][2]=sumlist[sno][2]+int(tlist[j+1])*int(tlist[-4])
            idx[sno]=idx[sno]+1
    for name in booklist:                
        sheetlist[name].row(0).write(0,u'客户内码')
        sheetlist[name].row(0).write(1,u'客户号')
        sheetlist[name].row(0).write(2,u'客户姓名')
        sheetlist[name].row(0).write(3,u'账号')
        sheetlist[name].row(0).write(4,u'6月30号余额')
        sheetlist[name].row(0).write(5,u'开户日期')
        sheetlist[name].row(0).write(6,u'销户日期')
        sheetlist[name].row(0).write(7,u'状态')
        sheetlist[name].row(0).write(8,u'账户类型')
        sheetlist[name].row(0).write(9,u'核心地址')
        sheetlist[name].row(0).write(10,u'信贷地址')
        sheetlist[name].row(0).write(11,u'存量日均')
        sheetlist[name].row(0).write(12,u'现量日均')
        sheetlist[name].row(0).write(13,u'增量日均')
        sheetlist[name].row(0).write(14,u'认定方式')
        sheetlist[name].row(0).write(15,u'客户经理号')
        sheetlist[name].row(0).write(16,u'管理类型')
        sheetlist[name].row(0).write(17,u'管理比例')
        sheetlist[name].row(0).write(18,u'所属网点')
        booklist[name].save(name+'存款.xls')
    sumbook = xlwt.Workbook()    
    sumsheet = sumbook.add_sheet('sheet1')   
    idx=1
    for sno in sumlist:
        tt=sno.split('_')
        tt.append("%.2f"%(sumlist[sno][0]/10000.0))
        tt.append("%.2f"%(sumlist[sno][1]/10000.0))
        tt.append("%.2f"%(sumlist[sno][2]/10000.0))
        #print tt
        for i in range(0,len(tt)):
            sumsheet.row(idx).write(i,str(tt[i]))
        idx=idx+1    
    sumbook.save("966110存款经理合计.xls")
def query_detail():
    try :
        db = util.DBConnect()
        q_sql="""
        select f.account_id,f.CST_NO,a.CST_ID,a.CST_NAME,a.ACCOUNT_NO,f.BALANCE,a.OPEN_DATE_ID,a.CLOSE_DATE_ID,s.ACCOUNT_STATUS,t.ACCOUNT_TYPE,i.CUST_ADDRESS,i.CUST_CREDIT_ADDRESS,f.YEAR_PDT / d.BEG_YEAR_DAYS,ah.FOLLOW_CUST,m.GROUP_KEY from F_BALANCE f
        join D_DATE d on f.DATE_ID=d.ID
        join D_ACCOUNT a on f.ACCOUNT_ID=a.ID
        join ACCOUNT_HOOK ah on ah.ACCOUNT_NO=a.ACCOUNT_NO
        join D_ACCOUNT_TYPE t on f.ACCOUNT_TYPE_ID=t.ID
        join D_ACCOUNT_STATUS s on f.ACCOUNT_STATUS_ID=s.ID
        left join D_CUST_INFO i on f.CST_NO=i.CUST_NO 
        join D_MANAGE m on f.MANAGE_ID=m.ID and m.MANAGE_TYPE<>'机构管理' and m.THIRD_BRANCH_CODE in ('966110','966111','966112','966113','966115')
        where f.DATE_ID=20160630 and f.ACCT_TYPE in (1,8)
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchall()
        print len(row)
        return row
    finally :
        db.closeDB()
def query_lrj():
    try :
        db = util.DBConnect()
        q_sql="""
        select f.ACCOUNT_ID,f.YEAR_PDT/ 365  from F_BALANCE f
        where f.DATE_ID=20151231 and f.ACCT_TYPE in (1,8) and f.YEAR_PDT>0
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchone()
        rowdict={}
        while row:
            rowdict[row[0]]=row[1]
            row=db.cursor.fetchone()
        print len(rowdict)    
        return rowdict
    finally :
        db.closeDB()

if __name__=='__main__':
    arglen=len(sys.argv)
    print "begin:", str(datetime.datetime.now())
    run_export()
    print "end:", str(datetime.datetime.now())
    #if arglen==3:
    #    filename=sys.argv[1]
    #    hooktype=sys.argv[2]
    #    print filename,hooktype
    #    print "begin:", str(datetime.datetime.now())
    #    run_export(filename,hooktype)
    #    print "end:", str(datetime.datetime.now())
    #else:
    #    print "please input python export_hook.py [filename] [hooktype]"
