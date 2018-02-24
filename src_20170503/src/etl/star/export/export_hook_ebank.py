# -*- coding:utf-8 -*- 
from nose.tools import eq_, raises, assert_true
from sqlalchemy.orm import aliased
import unittest
import time,os,random,sys
import DB2
from decimal import Decimal

import etl.base.util as util

import xlrd,xlwt

from openpyxl import Workbook,load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font,Color
import datetime

def run_export():
    data=query_detail()
    hook=query_hook()
    delcon=del_contract()
    idx={}
    typelist=['无','无','无','无','无','无','无']
    typedict={'个人网上银行':6,'企业网上银行':7,'手机银行':8,'支付宝卡通':9,'丰收e支付':10,'ETC':11,'支付宝快捷支付':12}
    sheetlist={}
    booklist={}
    rowlist={}
    for i in range(0,len(data)):
        x=data[i]
        if x[6] in delcon:
            if x[-2]+x[3] in delcon.get(x[6]):
                #print x[-2]+x[3] 
                continue
        if x[0]+x[3] not in rowlist:
            [manage_code,htype]=hook.get(x[0]+x[3],['无','无'])
            tx=manage_code+'!电子银行管户!100!'+str(x[3])+'!966000!966000'
            if x[6]=='企业网上银行' and manage_code=='无':
                #print x[6]
                continue
            tlist=list(x[:-3])+typelist+tx.split('!')+[htype]
            rowlist[x[0]+x[3]]=tlist
            rowlist[x[0]+x[3]][typedict[x[6]]]=x[6]
            if htype=='电子银行' and rowlist[x[0]+x[3]][7]=='无':
                rowlist[x[0]+x[3]][11]='ETC'
        else:
            rowlist[x[0]+x[3]][typedict[x[6]]]=x[6]
    print len(rowlist)        
    for ii in rowlist:
        tlist=rowlist[ii]    
        if len(tlist)!=20:
            print tlist
            continue
        sno=tlist[3]
        if sno not in booklist:
            booklist[sno]=xlwt.Workbook() 
            sheetlist[sno]=booklist[sno].add_sheet('sheet1')
            idx[sno]=1
        tlist.pop(17)
        tlist.pop(17)
        tlist.pop(15)
        temp=tlist[13]
        tlist[13]=tlist[14]
        tlist[14]=temp
        temp=tlist[14]
        tlist[14]=tlist[15]
        tlist[15]=temp
        temp=tlist[15]
        tlist[15]=tlist[16]
        tlist[16]=temp
        temp=tlist[14]
        tlist[14]=tlist[15]
        tlist[15]=temp
        for j in range(0,len(tlist)):
            sheetlist[sno].row(idx[sno]).write(j,str(tlist[j]).decode('utf-8'))
        idx[sno]=idx[sno]+1
    for name in booklist:                
        sheetlist[name].row(0).write(0,u'客户内码')
        sheetlist[name].row(0).write(1,u'客户号')
        sheetlist[name].row(0).write(2,u'客户姓名')
        sheetlist[name].row(0).write(3,u'签约机构')
        sheetlist[name].row(0).write(4,u'核心地址')
        sheetlist[name].row(0).write(5,u'信贷地址')
        sheetlist[name].row(0).write(6,u'个人网上银行')
        sheetlist[name].row(0).write(7,u'企业网上银行')
        sheetlist[name].row(0).write(8,u'手机银行')
        sheetlist[name].row(0).write(9,u'支付宝卡通')
        sheetlist[name].row(0).write(10,u'丰收e支付')
        sheetlist[name].row(0).write(11,u'ETC')
        sheetlist[name].row(0).write(12,u'支付宝快捷支付')
        sheetlist[name].row(0).write(13,u'管理类型')
        sheetlist[name].row(0).write(14,u'认定方式')
        sheetlist[name].row(0).write(15,u'所属网点')
        sheetlist[name].row(0).write(16,u'归属柜员号')
        if not os.path.exists('tmp/'+name[:-1]+'0/电子银行客户号/'):
            os.makedirs('tmp/'+name[:-1]+'0/电子银行客户号/')
        booklist[name].save('tmp/'+name[:-1]+'0/电子银行客户号/'+name+'电子银行客户汇总.xls')
        #booklist[name].save(name+'电子银行.xls')
    #sumbook = xlwt.Workbook()    
    #sumsheet = sumbook.add_sheet('sheet1')   
    #idx=1
    #for sno in sumlist:
    #    tt=sno.split('_')
    #    tt.append("%.2f"%(sumlist[sno][0]/10000.0))
    #    tt.append("%.2f"%(sumlist[sno][1]/10000.0))
    #    tt.append("%.2f"%(sumlist[sno][2]/10000.0))
    #    print tt
    #    for i in range(0,len(tt)):
    #        sumsheet.row(idx).write(i,str(tt[i]))
    #    idx=idx+1    
    #sumbook.save("966120电子银行经理合计.xls")
def query_detail():
    try :
        db = util.DBConnect()
        q_sql="""
        select trim(c.CST_NO),i.CUST_LONG_NO,i.CUST_NAME,o.org0_code,i.CUST_ADDRESS,i.CUST_CREDIT_ADDRESS,c.BUSI_TYPE,c.NET_CST_NO,m.GROUP_KEY from F_CONTRACT_STATUS f
        join D_CUST_CONTRACT c on f.CONTRACT_ID=c.ID and c.BUSI_TYPE in ('个人网上银行','企业网上银行','手机银行','支付宝卡通','丰收e支付','ETC','支付宝快捷支付') and left( CST_NO,1)='8' 
        left join D_CUST_INFO i on i.CUST_NO=c.CST_NO
        join D_MANAGE m on m.ID=f.MANAGE_ID
        join D_ORG o on o.id=f.org_id --and o.org1_code='966120'
        where f.DATE_ID=20160630 and f.status<>'不确定' 
        order by c.CST_NO,c.OPEN_BRANCH_NO,m.GROUP_KEY
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchall()
        q1_sql="""
        select trim(CST_NO),i.CUST_LONG_NO,i.CUST_NAME,case OPEN_BRANCH_NO when '966166' then '966163' else OPEN_BRANCH_NO end,i.CUST_ADDRESS,i.CUST_CREDIT_ADDRESS,BUSI_TYPE,NET_CST_NO,'无' from D_CUST_CONTRACT c 
        join D_CUST_INFO i on c.CST_NO=i.CUST_NO
        where BUSI_TYPE in ('支付宝卡通','支付宝快捷支付') --and OPEN_BRANCH_NO in ('966120','966121','966122','966123','966125')
        """
        db.cursor.execute(q1_sql)
        row1=db.cursor.fetchall()
        tlist=list(row)+list(row1)
        tlist.sort()
        print len(row),len(row1)
        return row+row1
    finally :
        db.closeDB()
def del_contract():
    try :
        db = util.DBConnect()
        sql="""
        select c.NET_CST_NO,c.OPEN_BRANCH_NO,c.BUSI_TYPE from F_CONTRACT_STATUS f
        join D_CUST_CONTRACT c on f.CONTRACT_ID=c.ID and left(c.OPEN_DATE,8)>f.DATE_ID and c.BUSI_TYPE in ('个人网上银行','手机银行','企业网上银行')
        where f.DATE_ID=20160630 and f.SUB_TYPE = '无'
        """
        db.cursor.execute(sql)
        row=db.cursor.fetchone()
        rowdict={}
        while row:
            if row[2] not in rowdict: 
                rowdict[row[2]]=[row[0]+row[1]]
            else:    
                rowdict[row[2]].append(row[0]+row[1])
            row=db.cursor.fetchone()
        print len(rowdict)    
        return rowdict
    finally :
        db.closeDB()
def query_hook():
    try :
        db = util.DBConnect()
        q_sql="""
        select distinct CUST_IN_NO,ORG_NO,MANAGER_NO from CUST_HOOK where TYP='存款' and HOOK_TYPE='管户'
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchone()
        rowdict={}
        while row:
            rowdict[row[0]+row[1]]=[row[2],'存款']
            row=db.cursor.fetchone()
        q_sql="""
        select distinct CUST_IN_NO,ORG_NO,MANAGER_NO from CUST_HOOK where TYP='贷款' and HOOK_TYPE='管户'
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchone()
        while row:
            rowdict[row[0]+row[1]]=[row[2],'贷款']
            row=db.cursor.fetchone()
        q_sql="""
        select distinct CUST_IN_NO,ORG_NO,MANAGER_NO from CUST_HOOK where TYP='电子银行' and HOOK_TYPE='管户'
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchone()
        while row:
            rowdict[row[0]+row[1]]=[row[2],'电子银行']
            row=db.cursor.fetchone()
        print len(rowdict)    
        return rowdict
    finally :
        db.closeDB()
if __name__=='__main__':
    arglen=len(sys.argv)
    print "begin:", str(datetime.datetime.now())
    run_export()
    print "end:", str(datetime.datetime.now())
    #if arglen==3:
    #    hooktype=sys.argv[2]
    #    print "end:", str(datetime.datetime.now())
    #else:
    #    print "please input python export_hook.py [filename] [hooktype]"
