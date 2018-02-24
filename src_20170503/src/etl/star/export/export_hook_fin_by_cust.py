# -*- coding:utf-8 -*- 
from nose.tools import eq_, raises, assert_true
from sqlalchemy.orm import aliased
import unittest
import time,os,random,sys
import DB2
from decimal import Decimal

import etl.base.util as util

import xlrd,xlwt

from openpyxl import Workbook,load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font,Color
import datetime

def run_export():
    dict2015=query_lrj()
    data=query_detail()
    hook=query_hook()
    idx={}
    sumlist={}
    sheetlist={}
    booklist={}
    rowlist={}
    for i in range(0,len(data)):
        temp=str(data[i][-1]).split(';')
        tempfc=str(data[i][-2])
        #print temp
        for x in temp:
            tt=[]
            tt.append(round(float(dict2015.get(data[i][0],0)/100.0),2))
            tt.append(round(float(data[i][12]/100.0),2))
            tt.append(0)
            tt.append(tempfc)
            #print tt
            tlist=list(data[i][:-3])+tt+x.split('!')
            tlist[5]=int(tlist[5])/100.0
            #else:    
            #if sno=="966115_9660518":print tlist
            tlist[14]=tlist[15]
            [tlist[16],tlist[15]]=hook.get(tlist[1]+tlist[19],['无','无'])
            if tlist[1]+tlist[19] not in rowlist:
                rowlist[tlist[1]+tlist[19]]=tlist
            else:
                rowlist[tlist[1]+tlist[19]][5]=rowlist[tlist[1]+tlist[19]][5]+tlist[5]
                rowlist[tlist[1]+tlist[19]][13]=rowlist[tlist[1]+tlist[19]][13]+tlist[13]
                rowlist[tlist[1]+tlist[19]][12]=rowlist[tlist[1]+tlist[19]][12]+tlist[12]
                #print rowlist[tlist[1]+tlist[19]][5],rowlist[tlist[1]+tlist[19]][13],rowlist[tlist[1]+tlist[19]][12]
    print len(rowlist)        
    for ii in rowlist:
        tlist=rowlist[ii]    
        sno=str(tlist[-3])
        if sno not in booklist:
            booklist[sno]=xlwt.Workbook() 
            sheetlist[sno]=booklist[sno].add_sheet('sheet1')
            idx[sno]=1
        tlist.pop(6)    
        tlist.pop(6)    
        tlist.pop(7)    
        tlist.pop(11)    
        tlist.pop(16)    
        tlist.pop(16)    
        temp=tlist[15]
        tlist[15]=tlist[14]
        tlist[14]=temp
        temp=tlist[13]
        tlist[13]=tlist[12]
        tlist[12]=temp
        temp=tlist[13]
        tlist[13]=tlist[14]
        tlist[14]=temp
        for j in range(0,len(tlist)-1):
            sheetlist[sno].row(idx[sno]).write(j,str(tlist[j+1]).decode('utf-8'))
        idx[sno]=idx[sno]+1
    for name in booklist:                
        sheetlist[name].row(0).write(0,u'客户内码')
        sheetlist[name].row(0).write(1,u'客户号')
        sheetlist[name].row(0).write(2,u'客户姓名')
        sheetlist[name].row(0).write(3,u'账号')
        sheetlist[name].row(0).write(4,u'购买金额')
        #sheetlist[name].row(0).write(5,u'开始日期')
        #sheetlist[name].row(0).write(6,u'到期日期')
        sheetlist[name].row(0).write(5,u'购买理财账户')
        #sheetlist[name].row(0).write(6,u'理财产品代码')
        sheetlist[name].row(0).write(6,u'核心地址')
        sheetlist[name].row(0).write(7,u'信贷地址')
        sheetlist[name].row(0).write(8,u'存量日均')
        sheetlist[name].row(0).write(9,u'现量日均')
        #sheetlist[name].row(0).write(10,u'购买渠道')
        sheetlist[name].row(0).write(10,u'认定方式')
        sheetlist[name].row(0).write(11,u'管理类型')
        sheetlist[name].row(0).write(12,u'所属网点')
        sheetlist[name].row(0).write(13,u'客户经理号')
        sheetlist[name].row(0).write(14,u'管理比例')
        if not os.path.exists('tmp/'+name[:-1]+'0/理财存量/'):
            os.makedirs('tmp/'+name[:-1]+'0/理财存量/')
        booklist[name].save('tmp/'+name[:-1]+'0/理财存量/'+name+'理财客户汇总.xls')
    #sumbook = xlwt.Workbook()    
    #sumsheet = sumbook.add_sheet('sheet1')   
    #idx=1
    #for sno in sumlist:
    #    tt=sno.split('_')
    #    tt.append("%.2f"%(sumlist[sno][0]/10000.0))
    #    tt.append("%.2f"%(sumlist[sno][1]/10000.0))
    #    tt.append("%.2f"%(sumlist[sno][2]/10000.0))
    #    print tt
    #    for i in range(0,len(tt)):
    #        sumsheet.row(idx).write(i,str(tt[i]))
    #    idx=idx+1    
    #sumbook.save("966110理财经理合计.xls")
def query_detail():
    try :
        db = util.DBConnect()
        q_sql="""
        select f.account_id,f.CST_NO,i.CUST_LONG_NO,i.CUST_NAME,a.FIN_ACCOUNT_NO,f.BALANCE,a.OPEN_DATE_ID,a.CLOSE_DATE_ID,a.DEP_ACCOUNT_NO,t.PROD_CODE,i.CUST_ADDRESS,i.CUST_CREDIT_ADDRESS,f.YEAR_PDT / 1.0/ d.BEG_YEAR_DAYS,a.CONTRACT_NO,'966110!理财管户!100!'||o.ORG0_CODE||'!966000!966000' from F_BALANCE f
        join D_DATE d on f.DATE_ID=d.ID
        join D_ACCOUNT a on f.ACCOUNT_ID=a.ID
        join D_ACCOUNT_TYPE t on f.ACCOUNT_TYPE_ID=t.ID
        join D_ACCOUNT_STATUS s on f.ACCOUNT_STATUS_ID=s.ID
        left join D_CUST_INFO i on f.CST_NO=i.CUST_NO 
        join D_MANAGE m on f.MANAGE_ID=m.ID -- and m.MANAGE_TYPE<>'机构管理' --and m.THIRD_BRANCH_CODE in ('966110','966111','966112','966113','966115')
        join D_ORG o on f.ORG_ID=o.ID --and o.ORG0_CODE in ('966110','966111','966112','966113','966115')
        where f.DATE_ID=20160630 and f.ACCT_TYPE=8
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchall()
        print len(row)
        return row
    finally :
        db.closeDB()
def query_lrj():
    try :
        db = util.DBConnect()
        q_sql="""
        select f.ACCOUNT_ID,f.YEAR_PDT/ 365.0  from F_BALANCE f
        where f.DATE_ID=20151231 and f.ACCT_TYPE=8 and f.YEAR_PDT>0
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchone()
        rowdict={}
        while row:
            rowdict[row[0]]=row[1]
            row=db.cursor.fetchone()
        print len(rowdict)    
        return rowdict
    finally :
        db.closeDB()
def query_hook():
    try :
        db = util.DBConnect()
        q_sql="""
        select distinct CUST_IN_NO,ORG_NO,MANAGER_NO from CUST_HOOK where TYP='存款' and HOOK_TYPE='管户'
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchone()
        rowdict={}
        while row:
            rowdict[row[0]+row[1]]=[row[2],'存款']
            row=db.cursor.fetchone()
        q_sql="""
        select distinct CUST_IN_NO,ORG_NO,MANAGER_NO from CUST_HOOK where TYP='贷款' and HOOK_TYPE='管户'
        """
        db.cursor.execute(q_sql)
        row=db.cursor.fetchone()
        while row:
            rowdict[row[0]+row[1]]=[row[2],'贷款']
            row=db.cursor.fetchone()
        print len(rowdict)    
        return rowdict
    finally :
        db.closeDB()
if __name__=='__main__':
    arglen=len(sys.argv)
    print "begin:", str(datetime.datetime.now())
    run_export()
    print "end:", str(datetime.datetime.now())
    #if arglen==3:
    #    filename=sys.argv[1]
    #    hooktype=sys.argv[2]
    #    print filename,hooktype
    #    print "begin:", str(datetime.datetime.now())
    #    run_export(filename,hooktype)
    #    print "end:", str(datetime.datetime.now())
    #else:
    #    print "please input python export_hook.py [filename] [hooktype]"
