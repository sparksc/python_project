# -*- coding:utf-8 -*- 
import os
#print os.getcwd()
import sys
#print sys.path
sys.path.append('/home/develop/src')
from nose.tools import eq_, raises, assert_true
from sqlalchemy.orm import aliased
import unittest
import time,os,random,sys
import DB2
from decimal import Decimal

import etl.base.util as util
import etl.star.update.update_hook_balance as uhb
import xlrd

from openpyxl import Workbook,load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font,Color
import datetime

def run_import(filename,org_no=None,subtype=None):
    #if subtype==None:subtype=hooktype
    data = xlrd.open_workbook(filename)
    filename='第三方存管'
    print filename
    sheet = data.sheet_by_index(0)
    nrows = sheet.nrows
    cust_sql="""
    INSERT INTO YDW.CUST_HOOK(CUST_NO,CUST_IN_NO,MANAGER_NO,ORG_NO,PERCENTAGE,HOOK_TYPE,START_DATE,END_DATE,STATUS,ETL_DATE,SRC,TYP,SUB_TYP,NOTE) 
    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """
    del_cust_sql="""
    DELETE FROM CUST_HOOK WHERE TYP='第三方存管' AND CUST_IN_NO= ? AND ORG_NO = ?
    """
    u_sql="""
    update F_CONTRACT_STATUS f set ORG_ID=(select ID from D_ORG where ORG0_CODE= ?)
    where f.CONTRACT_ID in (select c.ID from D_CUST_CONTRACT c where c.BUSI_TYPE= ? and c.NET_CST_NO= ?) and f.date_id=20160630
    """
    rc=[]
    ru=[]
    rd=[]
    indexdict={
        '第三方存管':['cust',0,17,16,0,-1],
    }
    custdict=util.get_cust_info()
    staffdict=util.get_staff_branch()
    for r in range(1,nrows):
        row = sheet.row_values(r)
        CUST_IN_NO=str(row[indexdict[filename][1]]).strip()
        MANAGER_NO=    row[indexdict[filename][2]]
        ORG_NO    =str(row[indexdict[filename][3]]).strip()
        #print CUST_IN_NO,MANAGER_NO,ORG_NO
        if MANAGER_NO=='' or ORG_NO=='':
            #print row,'1'
            continue    
        if CUST_IN_NO[-2:]=='.0':
            CUST_IN_NO=CUST_IN_NO[:-2]
        CUST_IN_NO=str(int(CUST_IN_NO))
        MANAGER_NO=str(int(MANAGER_NO)).strip()
        ORG_NO    =str(int(float(ORG_NO))).strip()
        #print MANAGER_NO,ORG_NO
        if MANAGER_NO[0:3]!='966' or len(ORG_NO)!=6 or ORG_NO[0:3]!='966' or len(MANAGER_NO)!=7:
            print row,'2'
            continue
        staffinfo=staffdict.get(MANAGER_NO)   
        if staffinfo is None:
            print MANAGER_NO,'!!!!'
            continue
        if staffinfo[1]=='非客户经理' and staffinfo[0][:-1]!=ORG_NO[:-1]:
            print ORG_NO,staffdict.get(MANAGER_NO)
            print '!!!!!',row
            continue
        if staffinfo[1]=='客户经理' and staffinfo[0]!=ORG_NO:
            print ORG_NO,staffdict.get(MANAGER_NO)
            print '!!!!!',row
            continue
        #if indexdict[filename][0]=='cust':
        #    CUST_IN_NO=str(int(sheet.cell(r,indexdict[filename][1]).value))
        if subtype=='ETC':CUST_IN_NO=str(util.get_cust_in_no_by_creditcard_no(str(row[indexdict[filename][1]])))
        if CUST_IN_NO=='0':print row,'3'
        #    #print CUST_IN_NO
        #elif indexdict[filename][0]=='account':
        #    CUST_IN_NO=str(int(sheet.cell(r,indexdict[filename][-1]).value))
        #    ACCOUNT_NO=str(int(sheet.cell(r,indexdict[filename][1]).value))
        HOOK_TYPE='管户'
        PERCENTAGE=100    
        START_DATE=20160630
        END_DATE=20991231
        STATUS='正常'
        ETL_DATE=20160630
        SRC='存量导入'
        TYP='第三方存管'
        SUB_TYP='第三方存管'
        info=custdict.get(CUST_IN_NO,('0','核心地址为:暂无','信贷地址为:暂无'))
        CUST_NO=info[0]
        if CUST_NO=='0':
            CUST_NO=str('101'+row[5].strip())
            print CUST_IN_NO,'4'
        if info[0]=='0':
            NOTE='地址暂无'
        if info[2] is None:
            NOTE='核心地址为:'+str(info[1])
        else:    
            NOTE='信贷地址为:'+str(info[2])
        rd.append((CUST_IN_NO,ORG_NO))
        #netkey=row[indexdict[filename][4]]
        #if type(netkey)==type(1.0):
        #    netkey=str(int(netkey))
        #else:
        #    netkey=netkey.strip()
        #if subtype=='ETC':
        #    netkey=netkey[1:]
        #ru.append((ORG_NO,subtype,str(netkey.encode('utf-8'))))
        if (CUST_NO,CUST_IN_NO,MANAGER_NO,ORG_NO,PERCENTAGE,HOOK_TYPE,START_DATE,END_DATE,STATUS,ETL_DATE,SRC,TYP,SUB_TYP,NOTE) not in rc:
            rc.append((CUST_NO,CUST_IN_NO,MANAGER_NO,ORG_NO,PERCENTAGE,HOOK_TYPE,START_DATE,END_DATE,STATUS,ETL_DATE,SRC,TYP,SUB_TYP,NOTE))
    print len(rd), str(datetime.datetime.now())
    insert_to_hook(del_cust_sql,rd)    
    print len(rc), str(datetime.datetime.now())
    insert_to_hook(cust_sql,rc)    
    #print len(ru), str(datetime.datetime.now())
    #insert_to_hook(u_sql,ru)    
        
def insert_to_hook(insert_sql,listdata):
    try :
        db = util.DBConnect()
        db.cursor.executemany(insert_sql,listdata)
        db.conn.commit()
    finally :
        db.closeDB()


if __name__=='__main__':
    arglen=len(sys.argv)
    if arglen==2:
        filename=sys.argv[1]
        print filename
        filelist=[x for x in os.listdir(filename) if os.path.splitext(x)[1].find('xls')]
        print "begin:", str(datetime.datetime.now())
        for filepath in filelist:
            if 'svn' not in filepath:
                print filepath
                run_import(filename+filepath)
        uhb.del_dup()
        print "end:", str(datetime.datetime.now())
    elif arglen==3:
        filename=sys.argv[1]
        orgcode=sys.argv[2]
        print "begin:", str(datetime.datetime.now())
        run_import(filename,orgcode)
        uhb.del_dup()
        print "end:", str(datetime.datetime.now())
    else:
        print "please input python import_hook.py [filename] [hooktype]"
