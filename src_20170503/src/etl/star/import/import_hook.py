# -*- coding:utf-8 -*- 
from nose.tools import eq_, raises, assert_true
from sqlalchemy.orm import aliased
import unittest
import time,os,random,sys
import DB2
from decimal import Decimal

import etl.base.util as util

import xlrd

from openpyxl import Workbook,load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font,Color
import datetime

def run_import(filename,hooktype,subtype=None):
    if subtype==None:subtype=hooktype
    data = xlrd.open_workbook(filename)
    sheet = data.sheet_by_index(1)
    nrows = sheet.nrows
    manBranDict = {}
    cust_sql="""
    INSERT INTO YDW.CUST_HOOK(CUST_NO,CUST_IN_NO,MANAGER_NO,ORG_NO,PERCENTAGE,HOOK_TYPE,START_DATE,END_DATE,STATUS,ETL_DATE,SRC,TYP,SUB_TYP,NOTE) 
    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """
    account_sql="""
    INSERT INTO YDW.ACCOUNT_HOOK(ACCOUNT_NO,MANAGER_NO,ORG_NO,PERCENTAGE,HOOK_TYPE,START_DATE,END_DATE,STATUS,ETL_DATE,SRC,TYP,SUB_TYP,NOTE) 
    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
    """
    rs=[]
    indexdict={
        '贷款户(已认).xls':['cust',1,12,6,-1,-1],
        '核销贷款认定.xls':['cust',1,36,0,-1,-1],
        '对公.xls':['cust',2,16,0,19,18],
        '对私存款.xls':['cust',2,16,0,-1,-1],
        '信用卡.xls':['account',0,14,11,-1,-1,2],
        '企业网银.xls':['cust',2,25,24,-1,-1],
        'ETC松门.xls':['cust',3,12,11,-1,-1],
    }
    if indexdict[filename][0]=='cust':
        i_sql=cust_sql        
    elif indexdict[filename][0]=='account':
        i_sql=account_sql
    custdict=util.get_cust_info()
    for r in range(1,nrows):
        row = sheet.row_values(r)
        if row[indexdict[filename][2]]=='' or row[indexdict[filename][3]]=='':
            print row
            continue    
        MANAGER_NO=str(int(sheet.cell(r,indexdict[filename][2]).value)).strip()
        ORG_NO=str(int(sheet.cell(r,indexdict[filename][3]).value)).strip()
        if len(MANAGER_NO)!=7 or MANAGER_NO[0:3]!='966' or len(ORG_NO)!=6 or ORG_NO[0:3]!='966':
            print row
            continue
        if indexdict[filename][0]=='cust':
            CUST_IN_NO=str(int(sheet.cell(r,indexdict[filename][1]).value))
            if subtype=='ETC':CUST_IN_NO=str(util.get_cust_in_no_by_creditcard_no(str(sheet.cell(r,indexdict[filename][1]).value)))
            #print CUST_IN_NO
        elif indexdict[filename][0]=='account':
            CUST_IN_NO=str(int(sheet.cell(r,indexdict[filename][-1]).value))
            ACCOUNT_NO=str(int(sheet.cell(r,indexdict[filename][1]).value))
        if indexdict[filename][4]>=0:
            PERCENTAGE=int(float(sheet.cell(r,indexdict[filename][4]).value)*100)
        else:
            PERCENTAGE=100    
        if indexdict[filename][5]>=0:
            if sheet.cell(r,indexdict[filename][5]).value==u'是':
                HOOK_TYPE='管户'
            else:    
                HOOK_TYPE='分润'
        else:
            HOOK_TYPE='管户'
        START_DATE=int(time.strftime("20%y%m%d"))
        END_DATE=20991231
        STATUS='正常'
        ETL_DATE=20160630
        SRC='存量导入'
        TYP=hooktype
        SUB_TYP=subtype
        info=custdict.get(CUST_IN_NO,('0','核心地址为:暂无','信贷地址为:暂无'))
        CUST_NO=info[0]
        NOTE=str(info[1])+str(info[2])
        if indexdict[filename][0]=='cust':
            rs.append((CUST_NO,CUST_IN_NO,MANAGER_NO,ORG_NO,PERCENTAGE,HOOK_TYPE,START_DATE,END_DATE,STATUS,ETL_DATE,SRC,TYP,SUB_TYP,NOTE))
        elif indexdict[filename][0]=='account':
            rs.append((ACCOUNT_NO,MANAGER_NO,ORG_NO,PERCENTAGE,HOOK_TYPE,START_DATE,END_DATE,STATUS,ETL_DATE,SRC,TYP,SUB_TYP,NOTE))
        #print rs[-1]    
    insert_to_hook(i_sql,rs)    
    print len(rs)
        
def insert_to_hook(insert_sql,listdata):
    try :
        db = util.DBConnect()
        db.cursor.executemany(insert_sql,listdata)
        db.conn.commit()
    finally :
        db.closeDB()


if __name__=='__main__':
    arglen=len(sys.argv)
    if arglen==3:
        filename=sys.argv[1]
        hooktype=sys.argv[2]
        print filename,hooktype
        print "begin:", str(datetime.datetime.now())
        run_import(filename,hooktype)
        print "end:", str(datetime.datetime.now())
    elif arglen==4:
        filename=sys.argv[1]
        hooktype=sys.argv[2]
        subtype=sys.argv[3]
        print filename,hooktype,subtype
        print "begin:", str(datetime.datetime.now())
        run_import(filename,hooktype,subtype)
        print "end:", str(datetime.datetime.now())
    else:
        print "please input python import_hook.py [filename] [hooktype]"
