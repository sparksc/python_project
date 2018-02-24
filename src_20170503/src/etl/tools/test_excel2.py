#! coding:utf-8
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

wb = Workbook()
ws = wb.create_sheet(title="test")

fp = open('/home/plwu/tmp/files/gbk.text','r')
dataListFile = []
while True:
    line = fp.readline()
    if not line:
        break
    line = line.replace('\"','')
    line = line.replace('\r','')
    line = line.replace('\n','')
    print type(line)
    g = line.decode('gbk')
    print type(g)
    u = g.encode('utf8')
    print type(u)
    value= u.split(',')
    dataListFile.append(value)

for r in dataListFile:
    ws.append(r)

#a1 = ws['A1']
#a2 = ws['A2']

ft = Font(color = colors.RED)
#a1.font = ft
#a2.font = ft

ws.column_dimensions['B'].width = 40
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 40
ws.column_dimensions['B'].font = ft
ws.column_dimensions['C'].font = ft
ws.column_dimensions['D'].font = ft

ws.row_dimensions[1].font = ft
ws.row_dimensions[2].font = ft
ws.row_dimensions[1].height= 60 
ws.row_dimensions[2].height= 60
wb.save("test.xlsx")
