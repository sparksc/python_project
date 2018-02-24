# -*- coding:utf-8 -*- 
from nose.tools import eq_, raises, assert_true
from sqlalchemy.orm import aliased
import unittest
import time
import DB2
from decimal import Decimal

import etl.base.util as util

import xlrd

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
import os
import datetime

from openpyxl.styles import colors
from openpyxl.styles import Font, Color

def result_to_file_jigou_excel(old_file_name, index, dList, head, col, dec):
    pathu = u'/home/plwu/tmp/files/deal/' + old_file_name + '/'
    path = pathu.encode('gbk')
    if os.path.exists(path) == False:
        os.makedirs(path)

    dList.sort(key=lambda x:x[index])    #开户网点号
    d_len = len(dList)
    i = 2
    org_no = dList[1][index].strip()

    wb = Workbook()
    ws = wb.active

    #wb = Workbook(write_only=True)
    #ws = wb.create_sheet()

    decs = dec.split(',')
    ws.append(head)
    dd = []
    j = 0
    for ddd in dList[1]:
        try:
            decs.index(str(j))
        except:
            dd.append(ddd)
            j = j + 1
            continue

        temp = Decimal(ddd)
        res = '{:,}'.format(temp)
        print temp, res
        dd.append(res)
        j = j + 1

    ws.append(dd)
    file_name_u = pathu + org_no.decode('utf8') + "_" + old_file_name+ '.xlsx'
    file_name = file_name_u.encode('gbk')

    while True:
        if i >= d_len:
            break;
        new_org_no = dList[i][index].strip().decode('utf8')
        if new_org_no != org_no:   #更换机构号
            if col != 'NO':
                cols = col.split(',')
                for cc in cols:
                    print cc
                    t = ws[str(cc)]
                    t.font = Font(color=colors.RED)
            wb.save(filename = file_name)

            wb = Workbook()
            ws = wb.active
            #wb = Workbook(write_only=True)
            #ws = wb.create_sheet()
            ws.append(head)
            file_name_u = pathu + new_org_no + "_" + old_file_name + '.xlsx'
            file_name = file_name_u.encode('gbk')

        try:
            dd = []
            j = 0
            for ddd in dList[i]:
                try:
                    decs.index(str(j))
                except:
                    dd.append(ddd)
                    j = j + 1
                    continue

                temp = Decimal(ddd)
                res = '{:,}'.format(temp)
                #print temp, res
                dd.append(res)
                j = j + 1
            ws.append(dd)
        except:
            print i, d_len

        i = i + 1
        org_no = new_org_no

    if col != 'NO':
        cols = col.split(',')
        for cc in cols:
            print cc, str(cc)
            t = ws[str(cc)]
            t.font = Font(color=colors.RED)

    wb.save(filename = file_name)


def read_file_to_list(file_name):
    print '/home/plwu/tmp/files/' + file_name
    fp = open('/home/plwu/tmp/files/' + file_name,'r')
    count = 0
    dataList = []
    head = []
    while True:
        line = fp.readline()
        if not line:
            break
        #line = line.replace('\"','')
        line = line.replace('\r','')
        line = line.replace('\n','')
        u = ""
        '''
        try:
            line_encoding = chardet.detect(line).get('encoding')
            print line_encoding
            g = line.decode(line_encoding)
            u = g.encode('utf8')
        except:
            print "error:", count
        '''
        import chardet
        line_encoding = chardet.detect(line).get('encoding')
        try:
            g = ""
            if not line_encoding:
                g = line.decode('gb18030')
            else:
                g = line.decode(line_encoding)
            u = g.encode('utf8')
        except Exception as e:
            print e,count
            continue 
        s_len = len(u)
        uu = ""
        tag = 0
        for j in range(0, s_len):
            if u[j] == '\"':
                tag = tag + 1
            if tag % 2 == 1:        #奇数
                if u[j] == ',':
                    uu = uu + '|'
                    continue
            uu = uu + u[j]
        uuu = uu.replace('\"','')
        value = uuu.split(',')
        if count == 0:
            head = value
        else:
            dataList.append(value)
        count = count + 1
    print count
    fp.close()
    return head, dataList

if __name__=='__main__':
    print "begin:", str(datetime.datetime.now())
    """
    文件名,分组列,标红列,金额列
    """
    files = [
        [u'ETC_2016_1_5.txt', '10', 'L1,M1,N1','']
        [u"wenling20160630_wanyishang.del", "0", 'NO', "6,7,8,9,10"]
    ]
    """
    files = [
        [u"1.txt", '0', 'AK1,AL1','NO', "0"],
    ]
    """
    for d in files:
        print d[0], int(d[1])
        str1 = d[0].split('.')
        print d[0], int(d[1]), str1[0]
        head, dataList = read_file_to_list(d[0])
        result_to_file_jigou_excel(str1[0], int(d[1]), dataList, head, d[2], d[3])

    print "end:", str(datetime.datetime.now())
