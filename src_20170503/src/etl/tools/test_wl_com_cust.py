# -*- coding:utf-8 -*- 
from nose.tools import eq_, raises, assert_true
from sqlalchemy.orm import aliased
import unittest
import time
import DB2
from decimal import Decimal

import etl.base.util as util

import xlrd

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font,Color
 

u'''  温岭客户划分 '''
def test_init_mx20160511_manager_relation_full():
    db = util.DBConnect()
    sql = """
    INSERT INTO MX20160511_MANAGER_RELATION_FULL(JIGOU, HUIJIHU, ACCOUNT, CUST_NO, MAN_ID, 
                ACCT_NAME, BIZHONG, RIJUN_2015, MONTH_RIJUN_201512, RIJUN_2016, 
                YM_BAL_201605, OPEN_DATE, XIAOHU, MANAGER_ID, FENRUN, 
                FLAG, ADDRESS1, ADDRESS2, TEL) 
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    fp = open('/home/plwu/tmp/files/mx20160511_manager_relation_full.csv.utf-8','r')
    count = 0
    dataList = []
    while True:
        line = fp.readline()
        if not line:
            break
        line = line.replace('\"','')
        line = line.replace('\r','')
        line = line.replace('\n','')
        line.decode('utf-8').encode('gbk')
        #print line
        value = line.split(',')
        #print value
        type(value)
        dataList.append(value)
        count = count + 1

    print count

    db.cursor.executemany(sql, dataList)
    db.conn.commit()

def test_init_mx20160511_manager_relation_full_one_by_one():
    db = util.DBConnect()
    sql = """
    INSERT INTO MX20160511_MANAGER_RELATION_FULL(JIGOU, HUIJIHU, ACCOUNT, CUST_NO, MAN_ID, 
                ACCT_NAME, BIZHONG, RIJUN_2015, MONTH_RIJUN_201512, RIJUN_2016, 
                YM_BAL_201605, OPEN_DATE, XIAOHU, MANAGER_ID, FENRUN, 
                FLAG, ADDRESS1, ADDRESS2, TEL) 
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    fp = open('/home/plwu/tmp/files/mx20160511_manager_relation_full.csv.utf-8','r')
    count = 0
    dataList = []
    while True:
        line = fp.readline()
        if not line:
            break
        line = line.replace('\"','')
        line = line.replace('\r','')
        line = line.replace('\n','')
        line.decode('utf-8').encode('gbk')
        print line
        value = line.split(',')
        print value
        db.cursor.execute(sql, value)
        count = count + 1

    print count
    db.conn.commit()

def result_to_file(file_name, dList):
    fp = open(file_name,'w')
    for l in dList:
        s = ','
        s = s.join(l)
        fp.write(s)
        fp.write('\n')
    fp.close() 

def result_to_file_jigou(flag, dList):
    path = '/home/plwu/tmp/files/result/'

    dList.sort(key=lambda x:x[0])    #开户网点号
    d_len = len(dList)
    i = 1
    org_no = dList[0][0].strip()
    file_name = path + "result_" + str(flag) + "_" + org_no
    fp = open(file_name,'w')
    s = ','
    s = s.join(dList[0])
    fp.write(s)
    fp.write('\n')

    while True:
        if i >= d_len:
            break;
        new_org_no = dList[i][0].strip()
        if new_org_no != org_no:   #更换机构号
            fp.close()
            file_name = path + "result_" + str(flag) + "_" + new_org_no
            fp = open(file_name,'w')

        s = ','
        s = s.join(dList[i])
        fp.write(s)
        fp.write('\n')
        i = i + 1
        org_no = new_org_no
    fp.close() 

def result_to_file_jigou_excel(flag, dList):
    path = '/home/plwu/tmp/files/excel/'

    dList.sort(key=lambda x:x[0])    #开户网点号
    d_len = len(dList)
    i = 1
    org_no = dList[0][0].strip()

    wb = Workbook()
    ws = wb.active
    head = [u"网点号",u"归集户",u"账号",u"客户号",u"客户内码",u"客户名称",u"币种",u"2015年日均",u"2015年12日均",u"2016年日均",u"余额",u"原分配客户经理",u"比例",u"标记",u"核心地址",u"信贷地址",u"电话",u"销户日期",u"挂钩开始日期"]
    ws.append(head)
    ws.append(dList[0])
    file_name = path + "result2_" + str(flag) + "_" + org_no + '.xlsx'

    while True:
        if i >= d_len:
            break;
        new_org_no = dList[i][0].strip()
        if new_org_no != org_no:   #更换机构号
            wb.save(filename = file_name)
            wb = Workbook()
            ws = wb.active
            ws.append(head)
            file_name = path + "result2_" + str(flag) + "_" + new_org_no + '.xlsx'

        try:
            ws.append(dList[i])
        except:
            print i, d_len
            print dList[i]

        i = i + 1
        org_no = new_org_no

    wb.save(filename = file_name)

"""
    newDataList = []            #正常
    newDataList1 = []           #开户机构和客户经理所属机构不同
    newDataList2 = []           #同一客户,同一开户网点,多个账号归属不同客户经理且客户经理所属网点和开户网点相同
    newDataList3 = []           #登记柜员号不存在，退休的
    newDataList7 = []           #内码是81000000000,且账号日均小于10000
    newDataList5 = []           #归属8位虚拟柜员的

    pin jie xin dai yuan
"""
def result_to_file_jigou_cust_no_excel(newDataList, newDataList1, newDataList2, newDataList3, newDataList5,cust_day_amt_dict_2015, cust_day_amt_dict_2016, cust_amt_dict):
    path = '/home/plwu/tmp/files/excel_cust_no/'

    loanFileDict = read_loan_file_to_list()

    allDataList = []
    for d in newDataList:
        cust_no = d[4].strip()
        if cust_no[0:2] != '82':
            continue
        org_no = d[0].strip()
        loan_pk = org_no + ":" + cust_no
        #d[19] = u' '
        d.append(u'正常')
        if loanFileDict.has_key(loan_pk) == False:       #no
            d.append(u' ')
            d.append(u' ')
        else:
            str11 = loanFileDict[loan_pk]
            str1 = str11.split(',')
            d.append(str1[0])
            d.append(str1[1])
        allDataList.append(d)

    for d in newDataList1:
        cust_no = d[4].strip()
        if cust_no[0:2] != '82':
            continue
        org_no = d[0].strip()
        loan_pk = org_no + ":" + cust_no
        #d[19] = u' '
        d.append(u'开户机构和客户经理所属机构不同')
        if loanFileDict.has_key(loan_pk) == False:       #no
            d.append(u' ')
            d.append(u' ')
        else:
            str11 = loanFileDict[loan_pk]
            str1 = str11.split(',')
            d.append(str1[0])
            d.append(str1[1])
        allDataList.append(d)

    for d in newDataList2:
        cust_no = d[4].strip()
        if cust_no[0:2] != '82':
            continue
        org_no = d[0].strip()
        loan_pk = org_no + ":" + cust_no
        d.append(u'同一客户同一开户网点多个账号归属不同客户经理且客户经理所属网点和开户网点相同')
        if loanFileDict.has_key(loan_pk) == False:       #no
            d.append(u' ')
            d.append(u' ')
        else:
            str11 = loanFileDict[loan_pk]
            str1 = str11.split(',')
            d.append(str1[0])
            d.append(str1[1])
        allDataList.append(d)

    for d in newDataList3:
        cust_no = d[4].strip()
        org_no = d[0].strip()
        if cust_no[0:2] != '82':
            continue
        loan_pk = org_no + ":" + cust_no
        d.append(u'登记柜员号不存在')
        if loanFileDict.has_key(loan_pk) == False:       #no
            d.append(u' ')
            d.append(u' ')
        else:
            str11 = loanFileDict[loan_pk]
            str1 = str11.split(',')
            d.append(str1[0])
            d.append(str1[1])
        allDataList.append(d)

    for d in newDataList5:
        cust_no = d[4].strip()
        org_no = d[0].strip()
        if cust_no[0:2] != '82':
            continue
        loan_pk = org_no + ":" + cust_no
        d.append(u'归属8位虚拟柜员的')
        if loanFileDict.has_key(loan_pk) == False:       #no
            d.append(u' ')
            d.append(u' ')
        else:
            str11 = loanFileDict[loan_pk]
            str1 = str11.split(',')
            d.append(str1[0])
            d.append(str1[1])
        allDataList.append(d)

    #begin deal
    allDataList.sort(key=lambda x:x[0])    #开户网点号
    d_len = len(allDataList)
    i = 1
    org_no = allDataList[0][0].strip()

    custOrgDict = {}
    wb = Workbook()
    ws = wb.active
    head = [u"网点号",u"客户号",u"客户内码",u"客户名称",u"2015年日均",u"2016年日均",u"余额",u"比例",u"核心地址",u"信贷地址",u"电话",u"挂钩生效日期",u"信贷经办人员",u"信贷经办人员姓名", u"原分配客户经理", u"错误原因", u"认定柜员号",u"认定人姓名","是否主办(是、否)", "分成比例"]
    ws.append(head)
    pk = allDataList[0][0].strip() + ':' + allDataList[0][4].strip()       #cust org pk
    custOrgDict[pk] = 1

    excelList = []
    excelList.append(allDataList[0][0])     #网点号
    excelList.append(allDataList[0][3])     #客户号
    excelList.append(allDataList[0][4])     #客户内码
    excelList.append(allDataList[0][5])     #客户名称
    excelList.append(cust_day_amt_dict_2015[pk])     #2015年日均
    excelList.append(cust_day_amt_dict_2016[pk])     #2016年日均
    excelList.append(cust_amt_dict[pk])              #余额
    excelList.append(allDataList[0][12])              #比例
    excelList.append(allDataList[0][14])              #核心地址
    excelList.append(allDataList[0][15])              #信贷地址
    excelList.append(allDataList[0][16])              #电话
    excelList.append(allDataList[0][18])              #挂钩生效日期
    excelList.append(allDataList[0][20])              #信贷经办客户经理
    excelList.append(allDataList[0][21])              #信贷经办客户经理
    excelList.append(allDataList[0][11])              #原分配
    excelList.append(allDataList[0][19])              #错误原因

    ws.append(excelList)
    file_name_u = path + org_no + u"_对公客户认定" + '.xlsx'
    file_name = file_name_u.encode('gbk')

    while True:
        if i >= d_len:
            break;
        new_org_no = allDataList[i][0].strip()
        if new_org_no != org_no:   #更换机构号
            q1 = ws['Q1']
            r1 = ws['R1']
            s1 = ws['S1']
            t1 = ws['T1']
            q1.font = Font(color=colors.RED)
            r1.font = Font(color=colors.RED)
            s1.font = Font(color=colors.RED)
            t1.font = Font(color=colors.RED)
            wb.save(filename = file_name)
            wb = Workbook()
            ws = wb.active
            ws.append(head)
            file_name_u = path + new_org_no + u"_对公客户认定" + '.xlsx'
            file_name = file_name_u.encode('gbk')

        pk = allDataList[i][0].strip() + ':' + allDataList[i][4].strip()        #cust org pk
        if custOrgDict.has_key(pk) == False:       #no
            excelList = []
            excelList.append(allDataList[i][0])     #网点号
            excelList.append(allDataList[i][3])     #客户号
            excelList.append(allDataList[i][4])     #客户内码
            excelList.append(allDataList[i][5])     #客户名称
            excelList.append(cust_day_amt_dict_2015[pk])     #2015年日均
            excelList.append(cust_day_amt_dict_2016[pk])     #2016年日均
            excelList.append(cust_amt_dict[pk])              #余额
            excelList.append(allDataList[i][12])              #比例
            excelList.append(allDataList[i][14])              #信贷地址
            excelList.append(allDataList[i][15])              #地址
            excelList.append(allDataList[i][16])              #电话
            excelList.append(allDataList[i][18])              #挂钩生效日期
            excelList.append(allDataList[i][20])              #信贷经办客户经理
            excelList.append(allDataList[i][21])              #信贷经办客户经理
            excelList.append(allDataList[i][11])              #原分配
            excelList.append(allDataList[i][19])              #错误原因
            ws.append(excelList)
            custOrgDict[pk] = 1
        i = i + 1
        org_no = new_org_no

    q1 = ws['Q1']
    r1 = ws['R1']
    s1 = ws['S1']
    t1 = ws['T1']
    q1.font = Font(color=colors.RED)
    r1.font = Font(color=colors.RED)
    s1.font = Font(color=colors.RED)
    t1.font = Font(color=colors.RED)
    wb.save(filename = file_name)


'''构造list处理文件
对于开户机构与客户经理所属机构不同的暂未处理 TBD'''
def test_init_mx20160511_manager_relation_full_file():
    fp = open('/home/plwu/tmp/files/mx20160511_manager_relation_full.csv.utf-8','r')
    count = 0
    dataList = []
    while True:
        line = fp.readline()
        if not line:
            break
        line = line.replace('\"','')
        line = line.replace('\r','')
        line = line.replace('\n','')
        value= line.split(',')
        dataList.append(value)
        count = count + 1
    print count

    newDataList1 = []           #归属网点不同，且客户优先
    newDataList2 = []           #归属网点不同, 且账号优先
    newDataList3 = []           #归属不同客户经理，但网点相同
    newDataList = []            #正常
    dataList.sort(key=lambda x:x[3])
    man_dict = deal_cust_file()
    d_len = len(dataList)
    i = 1
    cus_no = dataList[0][3].strip()
    tempDataList = []
    tempDataList.append(dataList[0])
    while True:
        if i >= d_len:
            break;
        new_cus_no = dataList[i][3].strip()
        if new_cus_no != cus_no:   #更换客户编号
            tempDataList.sort(key=lambda x:x[11])
            flag = 0        #默认正常数据
            is_same_branch = 0      #默认相同
            tempDataListLen = len(tempDataList)
            if tempDataListLen == 1:
                newDataList.append(tempDataList[0])         #只有1个账号正确
            else:
                old_man_id = tempDataList[0][11].strip()
                for j in range(1, tempDataListLen):
                    t_type = tempDataList[j][15].strip().decode('utf-8')
                    man_id = tempDataList[j][11].strip()
                    if t_type == u'账户优先' or t_type == u'汇集户优先':       #只要有账户优先则全部算账户优先
                        flag = 1
                    elif t_type == u'客户优先' and flag != 1:      #全部客户优先才算客户优先
                        flag = 2

                    if old_man_id == man_id:  #同一客户经理
                        pass
                    else:
                        if len(old_man_id) == 6 or len(man_id) == 6:    #有归属客户经理的，也有直接归属机构的 TBD
                            pass
                        elif man_dict[old_man_id] != man_dict[man_id]:   #客户经理归属网点不同
                            is_same_branch = 1;
                            break
                        else:                                          #归属不同客户经理但是网点相同
                            is_same_branch = 0;

                for tl in tempDataList:         #一个客户全部放入一类
                    if is_same_branch == 1:
                        if flag == 1:
                            newDataList1.append(tl)
                        elif flag == 2:
                            newDataList2.append(tl)
                        else:
                            print is_same_branch,flag,t_type
                            raise Exception('flag error!')
                    else:
                        if flag == 0:
                            newDataList.append(tl)
                        else:
                            newDataList3.append(tl)

            tempDataList = []
            tempDataList.append(dataList[i])
        else:
            tempDataList.append(dataList[i])
            pass
        cus_no = new_cus_no
        i = i + 1

    path = '/home/plwu/tmp/files/'
    result_to_file(path+'result.txt', newDataList)
    result_to_file(path+'result1.txt', newDataList1)
    result_to_file(path+'result2.txt', newDataList2)
    result_to_file(path+'result3.txt', newDataList3)


'''构造list处理文件
处理开户机构与客户经理所属机构不同的'''
def cal_cust_day_amt_2015(dataList):
    custDayAmt = {}
    dataList.sort(key=lambda x:x[4].strip())    #客户号
    cus_no = dataList[0][4].strip()
    tempDataList = []                      #先按客户
    tempDataList.append(dataList[0])
    for dl in dataList:
        new_cus_no = dl[4].strip()
        if new_cus_no != cus_no:   #更换客户编号
            tempDataList.sort(key=lambda x:x[0].strip())    #JIGOU
            jigou = tempDataList[0][0].strip()
            day_amt_2015_sum = 0
            for tdl in tempDataList:
                new_jigou = tdl[0].strip()

                day_amt_2015 = Decimal(tdl[7].strip())
                per = Decimal(tdl[12].strip())
                if new_jigou != jigou:   #更换机构
                    pk = jigou + ':'
                    pk = pk + cus_no 
                    custDayAmt[pk] = day_amt_2015_sum
                    #print pk, str(day_amt_2015_sum)
                    day_amt_2015_sum = day_amt_2015 * per
                else:
                    day_amt_2015_sum = day_amt_2015_sum + day_amt_2015 * per
                jigou = new_jigou

            pk = jigou + ':'
            pk = pk + cus_no 
            custDayAmt[pk] = day_amt_2015_sum
            #print pk, str(day_amt_2015_sum)

            tempDataList = []
            tempDataList.append(dl)
        else:
            tempDataList.append(dl)
        cus_no = new_cus_no

    tempDataList.sort(key=lambda x:x[0].strip())    #JIGOU
    jigou = tempDataList[0][0].strip()
    day_amt_2015_sum = 0
    for tdl in tempDataList:
        new_jigou = tdl[0].strip()
        day_amt_2015 = Decimal(tdl[7].strip())
        per = Decimal(tdl[12].strip())
        if new_jigou != jigou:   #更换机构
            pk = jigou + ':'
            pk = pk + cus_no 
            custDayAmt[pk] = day_amt_2015_sum
        #print pk, str(day_amt_2015_sum)
            day_amt_2015_sum = day_amt_2015 * per
        else:
            day_amt_2015_sum = day_amt_2015_sum + day_amt_2015 * per
            jigou = new_jigou

    pk = jigou + ':'
    pk = pk + cus_no 
    custDayAmt[pk] = day_amt_2015_sum

    return custDayAmt
    

'''构造list处理文件
处理开户机构与客户经理所属机构不同的'''
def cal_cust_day_amt_2016(dataList):
    custDayAmt = {}
    dataList.sort(key=lambda x:x[4].strip())    #客户号
    cus_no = dataList[0][4].strip()
    tempDataList = []                      #先按客户
    tempDataList.append(dataList[0])
    for dl in dataList:
        new_cus_no = dl[4].strip()
        if new_cus_no != cus_no:   #更换客户编号
            tempDataList.sort(key=lambda x:x[0].strip())    #JIGOU
            jigou = tempDataList[0][0].strip()
            day_amt_2016_sum = 0
            for tdl in tempDataList:
                new_jigou = tdl[0].strip()
                day_amt_2016 = Decimal(tdl[9].strip())
                per = Decimal(tdl[12].strip())
                if new_jigou != jigou:   #更换机构
                    pk = jigou + ':'
                    pk = pk + cus_no 
                    custDayAmt[pk] = day_amt_2016_sum
                    #print pk, str(day_amt_2016_sum)
                    day_amt_2016_sum = day_amt_2016 * per
                else:
                    day_amt_2016_sum = day_amt_2016_sum + day_amt_2016 * per
                jigou = new_jigou

            pk = jigou + ':'
            pk = pk + cus_no 
            custDayAmt[pk] = day_amt_2016_sum
            #print pk, str(day_amt_2016_sum)

            tempDataList = []
            tempDataList.append(dl)
        else:
            tempDataList.append(dl)
        cus_no = new_cus_no

    tempDataList.sort(key=lambda x:x[0].strip())    #JIGOU
    jigou = tempDataList[0][0].strip()
    day_amt_2016_sum = 0
    for tdl in tempDataList:
        new_jigou = tdl[0].strip()
        day_amt_2016 = Decimal(tdl[9].strip())
        per = Decimal(tdl[12].strip())
        if new_jigou != jigou:   #更换机构
            pk = jigou + ':'
            pk = pk + cus_no 
            custDayAmt[pk] = day_amt_2016_sum
            #print pk, str(day_amt_2016_sum)
            day_amt_2016_sum = day_amt_2016 * per
        else:
            day_amt_2016_sum = day_amt_2016_sum + day_amt_2016 * per
            jigou = new_jigou

    pk = jigou + ':'
    pk = pk + cus_no 
    custDayAmt[pk] = day_amt_2016_sum

    return custDayAmt

'''构造list处理文件
处理开户机构与客户经理所属机构不同的'''
def cal_cust_amt(dataList):
    custDayAmt = {}
    dataList.sort(key=lambda x:x[4].strip())    #客户号
    cus_no = dataList[0][4].strip()
    tempDataList = []                      #先按客户
    tempDataList.append(dataList[0])
    for dl in dataList:
        new_cus_no = dl[4].strip()
        if new_cus_no != cus_no:   #更换客户编号
            tempDataList.sort(key=lambda x:x[0].strip())    #JIGOU
            jigou = tempDataList[0][0].strip()
            amt_sum = 0
            for tdl in tempDataList:
                new_jigou = tdl[0].strip()
                amt = Decimal(tdl[10].strip())
                if new_jigou != jigou:   #更换机构
                    pk = jigou + ':'
                    pk = pk + cus_no 
                    custDayAmt[pk] = amt_sum
                    amt_sum = amt
                else:
                    amt_sum = amt_sum + amt
                jigou = new_jigou

            pk = jigou + ':'
            pk = pk + cus_no 
            custDayAmt[pk] = amt_sum
            #print pk, str(amt_sum)

            tempDataList = []
            tempDataList.append(dl)
        else:
            tempDataList.append(dl)
        cus_no = new_cus_no

    tempDataList.sort(key=lambda x:x[0].strip())    #JIGOU
    jigou = tempDataList[0][0].strip()
    amt_sum = 0
    for tdl in tempDataList:
        new_jigou = tdl[0].strip()
        amt = Decimal(tdl[10].strip())
        if new_jigou != jigou:   #更换机构
            pk = jigou + ':'
            pk = pk + cus_no 
            custDayAmt[pk] = amt_sum
            amt_sum = amt
        else:
            amt_sum = amt_sum + amt
            jigou = new_jigou

    pk = jigou + ':'
    pk = pk + cus_no 
    custDayAmt[pk] = amt_sum

    return custDayAmt

def read_loan_file_to_list():
    fp = open('/home/plwu/tmp/files/贷款存量.txt','r')
    count = 0
    loanDict = {}
    while True:
        line = fp.readline()
        if not line:
            break
        #line = line.replace('\"','')
        line = line.replace('\r','')
        line = line.replace('\n','')
        try:
            g = line.decode('gbk')
            u = g.encode('utf8')
        except:
            print count

        s_len = len(u)
        uu = "" 
        tag = 0
        for j in range(0, s_len):
            if u[j] == '\"':
                tag = tag + 1
            if tag % 2 == 1:        #奇数
                if u[j] == ',':
                    uu = uu + '|'
                    continue
            uu = uu + u[j]
        uuu = uu.replace('\"','')
        value = uuu.split(',')
        count = count + 1
        str1 = value[9].strip() + ',' + value[10].strip()
        str2 = value[6].strip() + ':' + value[1].strip()
        loanDict[str2] = str1
    print count
    fp.close()
    return loanDict 


def test_init_mx20160511_manager_relation_full_file2():
    fp = open('/home/plwu/tmp/files/wenling20160701_01.del','r')
    #fp = open('/home/plwu/tmp/files/wenling20160511_01.del','r')
    #fp = open('/home/plwu/tmp/files/ceshi.txt','r')
    count = 0
    dataListFile = []
    while True:
        line = fp.readline()
        if not line:
            break
        line = line.replace('\r','')
        line = line.replace('\n','')
        try:
            g = line.decode('gbk')
            u = g.encode('utf8')
        except:
            print count

        s_len = len(u)
        uu = "" 
        tag = 0
        for j in range(0, s_len):
            if u[j] == '\"':
                tag = tag + 1
            if tag % 2 == 1:        #奇数
                if u[j] == ',':
                    continue
            uu = uu + u[j]
        uuu = uu.replace('\"','')
        value = uuu.split(',')

        cus_no = value[4].strip()
        if cus_no[0:2] != '82':
            continue
        dataListFile.append(value)
        count = count + 1
    print count
    fp.close()

    man_dict = deal_cust_file_xls() #客户经理所属机构dict
    cust_day_amt_dict_2015 = cal_cust_day_amt_2015(dataListFile)
    cust_day_amt_dict_2016 = cal_cust_day_amt_2016(dataListFile)
    cust_amt_dict = cal_cust_amt(dataListFile)

    newDataList6 = []           #2015年客户网点日均小于10000,不再参与后续的判断
    newDataList = []            #正常
    newDataList1 = []           #开户机构和客户经理所属机构不同
    newDataList2 = []           #同一客户,同一开户网点,多个账号归属不同客户经理且客户经理所属网点和开户网点相同
    newDataList3 = []           #登记柜员号不存在，退休的
    newDataList4 = []           #内码是81000000000,且2016年账号年日均大于10000
    newDataList7 = []           #内码是81000000000,且2016年账号年日均小于10000
    newDataList5 = []           #归属8位虚拟柜员的
    dataList = []               #排除开户机构和客户经理所属机构不同的LIST
    line = 1
    for dl in dataListFile:
        jigou = dl[0].strip()
        man_id = dl[11].strip()
        cus_no = dl[4].strip()
        line = line + 1
        if cus_no[0:2] != '82':
            continue

        ''' 账号日均>=10000 '''
        if cus_no == '81000000000':
            day_amt_2016 = Decimal(dl[9].strip())
            if day_amt_2016 < Decimal(10000.00):
                newDataList7.append(dl)
                continue
        else:
            pk = jigou + ':'
            pk = pk + cus_no 
            '''对公的不判断1W以下
            if cust_day_amt_dict_2015[pk] < Decimal(10000.00) and cust_day_amt_dict_2016[pk] < Decimal(10000.00):
                newDataList6.append(dl)
                continue
             '''
        if len(man_id) == 6:    #有归属客户经理的，也有直接归属机构的算正常
            dataList.append(dl)
            pass
        elif len(man_id) == 8:    #归属虚拟柜员的直接不再进行检查
            newDataList5.append(dl)
            pass
        else:
            if man_dict.has_key(man_id) == False:       #退休
                newDataList3.append(dl)
            else:
                if jigou != man_dict[man_id]:   #客户经理归属网点不同
                    newDataList1.append(dl)           #开户机构和客户经理所属机构不同
                else:
                    if cus_no == '81000000000':
                        newDataList4.append(dl)
                    else:
                        dataList.append(dl)

    print len(newDataList1),len(dataList)

    dataList.sort(key=lambda x:x[3])    #客户号
    d_len = len(dataList)
    i = 1
    cus_no = dataList[0][4].strip()
    tempDataList = []
    tempDataList.append(dataList[0])
    while True:
        if i >= d_len:
            break;
        new_cus_no = dataList[i][4].strip()
        if new_cus_no != cus_no:   #更换客户编号
            tempDataList.sort(key=lambda x:x[0])
            is_same_branch = 1      #默认相同
            tempDataListLen = len(tempDataList)
            if tempDataListLen == 1:
                newDataList.append(tempDataList[0])         #只有1个账号正确
            else:
                old_jigou_id = tempDataList[0][0].strip()
                for j in range(1, tempDataListLen):
                    jigou_id = tempDataList[j][0].strip()

                    if jigou_id != old_jigou_id:
                        is_same_branch = 0
                        break

                is_same_man = 1        #默认客户经理全部一样
                if is_same_branch == 1:
                    old_man_id = tempDataList[0][11].strip()
                    tempDataList.sort(key=lambda x:x[11])
                    for j in range(1, tempDataListLen):
                        man_id = tempDataList[j][11].strip()
                        if old_man_id == man_id:  #同一客户经理
                            pass
                        else:
                            if len(old_man_id) != len(man_id):    #有归属客户经理的，也有直接归属机构的,算归属不同人员
                                is_same_man = 0
                                break
                            else:
                                if old_man_id != man_id:   #客户经理归属不同
                                    is_same_man = 0
                                    break

                for tl in tempDataList:         #一个客户全部放入一类
                    if is_same_branch == 1:
                        if is_same_man == 0:
                            newDataList2.append(tl)
                        else:
                            newDataList.append(tl)
                    else:
                        newDataList.append(tl)

            tempDataList = []
            tempDataList.append(dataList[i])
        else:
            tempDataList.append(dataList[i])
            pass
        cus_no = new_cus_no
        i = i + 1

    """
    result_to_file_jigou_excel(0, newDataList)
    result_to_file_jigou_excel(1, newDataList1)
    result_to_file_jigou_excel(2, newDataList2)
    result_to_file_jigou_excel(3, newDataList3)
    result_to_file_jigou_excel(5, newDataList5)
    result_to_file_jigou_excel(6, newDataList6)
    """

    print "1:", len(newDataList1)
    print "2:", len(newDataList2)
    print "3:", len(newDataList3)
    print "5:", len(newDataList5)
    print "6:", len(newDataList6)

    result_to_file_jigou_cust_no_excel(newDataList, newDataList1, newDataList2, newDataList3, newDataList5, cust_day_amt_dict_2015, cust_day_amt_dict_2016, cust_amt_dict)


    '''
    result_to_file_jigou(0, newDataList)
    result_to_file_jigou(1, newDataList1)
    result_to_file_jigou(2, newDataList2)
    result_to_file_jigou(3, newDataList3)
    result_to_file_jigou(5, newDataList5)
    result_to_file_jigou(6, newDataList6)
    result_to_file_jigou(7, newDataList7)
    result_to_file_jigou(8, newDataList7)
    '''

def deal_cust_file():
    fp = open('/home/plwu/tmp/files/manager_branch.csv','r')

    count = 0
    manBranDict = {}
    while True:
        line = fp.readline()
        if not line:
            break
        value = line.split(',')
        manBranDict[value[0].strip()] = value[1].strip()
        count = count + 1
    return manBranDict
    fp.close()

def deal_cust_file_xls():
    data = xlrd.open_workbook('/home/plwu/tmp/files/柜员表.xls')
    sheet = data.sheet_by_index(0)
    nrows = sheet.nrows
    manBranDict = {}
    for r in range(2,nrows):
        #print sheet.cell(r,0).value, sheet.cell(r,2).value
        manBranDict[sheet.cell(r,0).value] = sheet.cell(r,2).value
    return manBranDict

if __name__=='__main__':
    test_init_mx20160511_manager_relation_full_file2()
    #deal_cust_file()
