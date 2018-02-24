# -*- coding:utf-8 -*-

import re
import json
from openpyxl import Workbook,load_workbook
from openpyxl.cell import column_index_from_string
from openpyxl.styles import Style, DEFAULTS as DEFAULTS_STYLE
import os

def excel_writer(template_name, data,to_excel):
    path = os.path.dirname(os.path.abspath(__file__))
    template_file  = os.path.join(path, "rpt_templates", template_name)
    wb = load_workbook(template_file)
    ws = wb.active
    exps = None
    for row in ws.rows:
        for cell in row:
            if cell and (isinstance(cell.value, (str, unicode))):
                idx = cell.value.find("exp:")
                if idx >= 0:
                    exps = cell.value[idx+4:]
                    break
    if exps is None:
        return  False

    exps = '{"start_row":2,"start_col":2}'
    exp = json.loads(exps)
    ridx = exp["start_row"]
    rcol = exp["start_col"]
    rows = data.get("rows")
    for ri,row in enumerate(rows, start=0):
        ri = ri + ridx 
        print ri
        col = rcol
        for v in row:
            cell = ws.cell(column = col, row = ri ,value=v)
            col = col + 1
    wb.save(to_excel)
    return  True

if __name__ == "__main__":
    data = {"rows":[["xx1","xxx2",3,4,5,6,7,8,9,10],["yy2",3,4,5,6,7,8,9,10,11],["yy2",3,4,5,6,7,8,9,10,11]],"header":["C1","C2","c3", "c4", "c5", "c6", "c7", "c8", "c9", "c10"]}
    excel_writer("000001.xlsx", data, "empty_book1.xlsx")
