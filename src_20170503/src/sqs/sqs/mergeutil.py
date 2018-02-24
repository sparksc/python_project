# -*- coding:utf-8 -*-
from openpyxl.styles import Font
from flask import current_app
class merge_excel:
    def __init__(self, ws, merge_style, exps, sub, cancalMerge=None):
        self.ws = ws
        self.merge_style = merge_style
        self.exps = exps
        self.sub = sub
        self.cancalMerge=cancalMerge
    def merge(self):
        ridx = self.exps["start_row"]
        rcol = self.exps["start_col"]
        if rcol <= 0:
            rcol = 1
        if ridx <= 0:
            ridx = 1
        cols = self.exps["cols"]
        font = Font( name=u'宋体',bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
        for ms in self.merge_style:
            if ms.find('-') != -1:
                if self.cancalMerge ==True:
                    current_app.logger.debug(type(self.cancalMerge))
                    continue
                #列合并起始行和结束行
                e_mgr_st = self.merge_style[ms][0] + ridx
                e_mgr_ed = self.merge_style[ms][1] + ridx
                #合并的列
                e_mgc = int(ms.split('-')[1]) + rcol
                self.ws.merge_cells( start_row=e_mgr_st, end_row=e_mgr_ed, start_column=e_mgc, end_column=e_mgc)
                #同含义列合并
                if self.sub is not None:
                    current_app.logger.debug(self.sub)
                    if self.sub.has_key(str(e_mgc-rcol)):
                        current_app.logger.debug(self.sub[str(e_mgc-rcol)])
                        for s_mgc in self.sub[str(e_mgc-rcol)]:
                            s_mgc += rcol
                            self.ws.merge_cells( start_row=e_mgr_st, end_row=e_mgr_ed, start_column=s_mgc, end_column=s_mgc)
            else:
                e_mgc_st = int(self.merge_style[ms][0]) + rcol
                e_mgc_ed = int(self.merge_style[ms][1]) + rcol
                e_mgr = int(ms)+ ridx
                self.ws.merge_cells( start_row=e_mgr, end_row=e_mgr, start_column=e_mgc_st, end_column=e_mgc_ed)
                self.ws.cell( column=e_mgc_st , row=e_mgr ).font = font
                
