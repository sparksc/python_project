# -*- coding:utf-8 -*-

import re
import json
from openpyxl import Workbook,load_workbook
from openpyxl.cell import column_index_from_string
from openpyxl.styles import Style, DEFAULTS as DEFAULTS_STYLE
import os
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import Alignment
from flask import current_app
from mergeutil import merge_excel
def excel_writer(template_name, data,to_excel):
    path = os.path.dirname(os.path.abspath(__file__))
    template_file  = os.path.join(path, "rpt_templates", template_name)
    wb = load_workbook(template_file)
    ws = wb.active
    exps = data.get("exps")

    if exps is None:
        for row in ws.rows:
            for cell in row:
                if cell and (isinstance(cell.value, (str, unicode))):
                    idx = cell.value.find("exp:")
                    if idx >= 0:
                        exps = cell.value[idx+4:]
                        break
            if exps is None: 
                return  False
            else:
                exps = json.loads(exps)
                break
    merge_style = data.get("merge_style")
    ridx = exps.get("start_row")
    rcol = exps.get("start_col")
    rows = data.get("rows")
    cols = exps.get("cols")
    styles = []
    load_style = False
    #print merge_style
    
    styles = Side(border_style="thin", color="000000")
    border = Border( left=styles, right=styles, top=styles, bottom=styles )
    alignment = Alignment( horizontal = 'center', vertical = 'center')
    font = Font( name=u'宋体',bold=False, italic=False, vertAlign=None, underline='none', strike=False)
    if ridx <=0 : ridx = 1
    for ri,row in enumerate(rows, start=0):
        ri = ri + ridx 
        if rcol <= 0 : rcol = 1
        idx = 0 
        for v in row:
            if v == 'None':
                v=''
            if isinstance(v,(str,unicode)):
                v = v.strip()
            cell = ws.cell(column = idx+rcol, row = ri ,value=v)
            cell.border = border 
            cell.alignment = alignment
            cell.font=font
            '''
            if load_style :
                cell.style = styles[idx]   
            else:
                styles.append( cell.style )
            '''
            idx = idx + 1
            if cols is not None and idx>=cols:
                break
        load_style = True

    current_app.logger.debug("xxxxxxxxxxxxxxxxxxxxxxxxxxx") 

    #合并分组
    if merge_style is not None:
        current_app.logger.debug(merge_style)
        sub = data.get("sub")
        cancalMerge=data.get("cancalMerge")
        current_app.logger.debug(sub)
        m = merge_excel(ws,merge_style, exps, sub, cancalMerge)
        m.merge()
    wb.save(to_excel)
    return  True
    
if __name__ == "__main__":
    data = {"rows":[["xx1","xxx2",3,4,5,6,7,8,9,10],["yy2",3,4,5,6,7,8,9,10,11],["yy2",3,4,5,6,7,8,9,10,11]],"header":["C1","C2","c3", "c4", "c5", "c6", "c7", "c8", "c9", "c10"]}
    excel_writer("000001.xlsx", data, "empty_book1.xlsx")
