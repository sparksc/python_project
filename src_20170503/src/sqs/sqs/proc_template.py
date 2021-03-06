# -*- coding:utf-8 -*-

import re
from openpyxl import Workbook,load_workbook
from openpyxl.cell import column_index_from_string
from openpyxl.styles import Style, DEFAULTS as DEFAULTS_STYLE
from openpyxl.styles import Side,Border,Alignment,Font

from mergeutil import merge_excel
import os
from flask import current_app

class TemplateProc(object):
        

    """Excel 模板处理"""
    def __init__(self, template_name=None):
        """TODO: to be defined1. """
        self.template_name = template_name
        self.template_file = None
        self.template_info = []
        self.template_re = re.compile("\{\{(?P<type>[A-Za-z]+)(?P<idx>\d+)\}\}")
        self.wb = None
        self.ws = None
        self._load_wb()
        self._get_template_info()

    def _load_wb(self):
        if self._template_file_exists():
           self.wb = load_workbook(self.template_file)

    def _template_file_exists(self):
        path = os.path.dirname(os.path.abspath(__file__))
        self.template_file  = os.path.join(path,"rpt_templates",self.template_name)
        return os.path.isfile(self.template_file)

    def _get_template_info(self):
        i = 1
        if self.wb:
            self.ws = self.wb.active
            for row in self.ws.rows:
                for cell in row:
                    if cell and (isinstance(cell.value, (str, unicode))):
                        m = self.template_re.match(cell.value)
                        if m:
                            #print m.group("idx"), m.group(1)
                            self.template_info.append({"dest_col":column_index_from_string(cell.column), \
                                    "dest_row":cell.row, \
                                    "src_col":int(m.group("idx"))-1, \
                                    "dest_col_style":self.ws.column_dimensions[cell.column].style, \
                                    "dest_type":m.group("type"), \
                                    "has_col_style":self.ws.column_dimensions[cell.column].style == DEFAULTS_STYLE})
                            i = i + 1

        print "style:", str(i)

    def get_excel(self, data):
        """SQS format data"""
        styles = Side(border_style="thin", color="000000")
        border = Border( left=styles, right=styles, top=styles, bottom=styles )
        alignment = Alignment( horizontal = 'center', vertical = 'center')
        font = Font( name=u'宋体',bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
        rows = data.get("rows")
        headers = data.get("header")
        for ti in self.template_info:
            src_col = ti.get("src_col")
            dest_type = ti.get("dest_type")
            dest_col = ti.get("dest_col")
            dest_row = ti.get("dest_row")
            dest_col_style = ti.get("dest_col_style")
            if src_col < len(headers): 
                if dest_type and dest_type.upper() == "H":
                    _ = self.ws.cell(column = dest_col,row = dest_row,value=headers[src_col])
                if dest_type and dest_type.upper() == "C":
                    for ri,row in enumerate(rows,start=0):
                        a1 = re.compile('[\x00-\x09\x0b-\x1f\x7f]')
                        temp = a1.sub('', str(row[src_col]))
                        if temp == 'None':
                            temp=''
                        if isinstance(temp,(str,unicode)):
                            temp=temp.strip()
                        cell = self.ws.cell(column = dest_col,row = dest_row+ri,value=temp)
                        cell.border = border
                        cell.alignment = alignment
                        cell.font = font
                        #self.merge_style(cell,dest_col_style)
                    #col_name = chr(ord(str(src_col)) + 17)
                    #self.ws.column_dimensions[col_name].style = dest_col_style
            else:
                _ = self.ws.cell(column = dest_col,row = dest_row,value="")
        #为总标题加粗，16号，宋体
        font_header = Font( name=u'宋体',size=16,bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
        #合并标题单元格
        '''
        for i in range(2,len(headers)+1):
            try:
                self.ws.unmerge_cells(start_row=1, end_row=1, start_column=1, end_column=i)
                break
            except:
                pass
        self.ws.merge_cells( start_row=1, end_row=1, start_column=1, end_column=len(headers))
        '''
        self.ws.cell(column=1,row=1).font = font_header
        self.ws.cell(column=1,row=1).alignment = alignment
        #为表头加黑框
        for i in range(len(headers)):
            cell=self.ws.cell(column=i+1,row=2)
            cell.font=font
            cell.alignment = alignment
            cell.border=border
        if data.get("merge_style") is not None:
                exps = {}
                exps["start_row"] = dest_row
                exps["start_col"] = 0
                exps["cols"] = len(headers)
                sub=data.get("sub")
                cancalMerge=data.get("cancalMerge")
                m = merge_excel(self.ws,data.get("merge_style"),exps,sub,cancalMerge)
                m.merge()
    def merge_style(self,cell,style):
        cell.style = style
        """
        if hash(cell.style.number_format) == hash(DEFAULTS_STYLE.number_format):
            cell.style.number_format = style.number_format
        if hash(cell.style.font) == hash(DEFAULTS_STYLE.font):
            cell.style.font = style.font
        if hash(cell.style.fill) == hash(DEFAULTS_STYLE.fill):
            cell.style.fill = style.fill
        # if hash(cell.style.borders) == hash(DEFAULTS_STYLE.borders):
        #     cell.style.borders = style.borders
        if hash(cell.style.alignment) == hash(DEFAULTS_STYLE.alignment):
            cell.style.alignment = style.alignment
        if hash(cell.style.protection) == hash(DEFAULTS_STYLE.protection):
            cell.style.protection = style.protection
        """

        
if __name__ == "__main__":
    tp = TemplateProc(template_name="000001.xlsx")
    print tp.template_file
    for i in tp.template_info:
        print i
    
    tp.get_excel({"rows":[[1,2,3,4,5,6,7,8,9,10]],"header":["C1","C2","c3", "c4", "c5", "c6", "c7", "c8", "c9", "c10"]})
    tp.wb.save('empty_book1.xlsx')
